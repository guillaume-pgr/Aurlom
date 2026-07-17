"""Lecture, validation et dérivation du classeur Excel (source du mode réel).

Contrat du fichier (cf. README) : 1 onglet = 1 table plate, en-têtes en ligne 1,
noms d'onglets et de colonnes figés. Charges toujours en négatif, pourcentages
en nombre (52 = 52 %), montants en k€ sauf colonnes suffixées ``_meur``.

Ce module ne fait QUE :
  1. parser le classeur en mémoire (openpyxl, read_only) ;
  2. le valider strictement (onglets, colonnes, types, 24 mois consécutifs,
     scénarios complets, mêmes campus_id partout, cohérences recalculées) ;
  3. le transformer en dict aligné sur get_dashboard_data().

Il n'écrit rien : le stockage est la responsabilité de excel_store.py.

Règle d'or : le dashboard CALCULE (totaux, cumuls, marges, croissances,
remplissage…), le fichier ne fournit que des valeurs saisies.
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import io
import re as _re
from decimal import ROUND_HALF_UP, Decimal

import openpyxl

# Tolérance des contrôles de cohérence recalculés (en k€).
TOLERANCE_KEUR = 1

# Onglet -> colonnes attendues, dans l'ordre exact.
SCHEMA: dict[str, list[str]] = {
    "parametres": ["cle", "valeur"],
    "mensuel_groupe": ["mois", "ca_reel_keur", "ca_budget_keur", "ebitda_reel_keur",
                       "ebitda_budget_keur", "tresorerie_brute_meur", "emprunts_meur",
                       "creances_opco_meur", "dso_opco_jours"],
    "pnl_ltm": ["poste", "montant_keur", "montant_n1_keur"],
    "tresorerie_13s": ["semaine", "solde_reel_meur", "solde_projete_meur"],
    "covenants": ["nom", "valeur", "seuil", "sens"],
    "echeances_lbo": ["libelle", "date", "montant_keur"],
    "inscriptions": ["mois", "rentree_n", "rentree_n1"],
    "filieres_groupe": ["filiere", "etudiants"],
    "acquisition": ["indicateur", "valeur", "variation_n1"],
    "campus": ["campus_id", "nom", "adresse", "capacite", "etudiants", "pct_alternance",
               "reussite_bts_pct", "effectif_moyen_classe", "attrition_pct", "nps",
               "cout_prof_keur", "loyer_annuel_keur", "surface_m2"],
    "campus_mensuel": ["campus_id", "mois", "ca_keur"],
    "campus_pnl": ["campus_id", "ca_alternance", "ca_initiale", "ca_stages_prepas",
                   "couts_pedagogiques", "masse_salariale_support", "loyers",
                   "marketing_admissions", "autres_opex"],
    "previsionnel": ["scenario", "mois", "ca_keur", "ebitda_keur", "tresorerie_meur"],
    "hypotheses": ["scenario", "ordre", "texte"],
    "points_cles": ["numero", "mois", "texte"],
}

# Onglets facultatifs : importés s'ils sont présents, ignorés sinon.
# - campus_filieres : répartition par filière AU NIVEAU CAMPUS. Sans lui, la carte
#   « Étudiants par filière » du détail campus est masquée (on n'invente rien).
OPTIONAL_SCHEMA: dict[str, list[str]] = {
    "campus_filieres": ["campus_id", "filiere", "etudiants"],
}

# Onglets présents dans le classeur mais explicitement ignorés à l'import :
# - controles : contrôles de cohérence -> on les RECALCULE en Python, on ne les lit pas.
# - LISEZ-MOI : notice destinée au rédacteur du fichier.
IGNORED_SHEETS = {"controles", "LISEZ-MOI"}

# Paramètres obligatoires de l'onglet 'parametres'.
REQUIRED_PARAMS = ["exercice_libelle", "ltm_debut", "ltm_fin", "date_maj"]

# Postes du P&L groupe : 3 produits (CA) puis 5 charges (négatives).
PNL_CA_POSTES = ["ca_alternance_opco", "ca_scolarite_initiale", "ca_stages_prepas"]
PNL_CHARGE_POSTES = ["couts_pedagogiques", "masse_salariale_support", "loyers_campus",
                     "marketing_admissions", "autres_opex"]

# Colonnes de campus_pnl : 3 produits puis 5 charges.
CAMPUS_CA_COLS = ["ca_alternance", "ca_initiale", "ca_stages_prepas"]
CAMPUS_CHARGE_COLS = ["couts_pedagogiques", "masse_salariale_support", "loyers",
                      "marketing_admissions", "autres_opex"]

# Scénarios du prévisionnel : clé fichier -> clé interne (celle du template).
SCENARIOS = {"base": "base", "bas": "low", "haut": "high"}
SCENARIO_LABELS = {"base": "Base", "low": "Rentrée −10 %", "high": "Rentrée +10 %"}

# Libellés affichés du P&L groupe (l'ordre suit PNL_CA_POSTES + PNL_CHARGE_POSTES).
PNL_LABELS = {
    "ca_alternance_opco": "dont alternance (OPCO)",
    "ca_scolarite_initiale": "dont scolarité initiale",
    "ca_stages_prepas": "dont stages & prépas",
    "couts_pedagogiques": "Coûts pédagogiques",
    "masse_salariale_support": "Masse salariale support",
    "loyers_campus": "Loyers campus",
    "marketing_admissions": "Marketing & admissions",
    "autres_opex": "Autres opex",
}
CAMPUS_PNL_LABELS = {
    "ca_alternance": "dont alternance (OPCO)",
    "ca_initiale": "dont scolarité initiale",
    "ca_stages_prepas": "dont stages & prépas",
    "couts_pedagogiques": "Coûts pédagogiques",
    "masse_salariale_support": "Masse salariale support",
    "loyers": "Loyers",
    "marketing_admissions": "Marketing & admissions",
    "autres_opex": "Autres opex",
}

_MOIS_COURTS = {1: "janv.", 2: "févr.", 3: "mars", 4: "avr.", 5: "mai", 6: "juin",
                7: "juil.", 8: "août", 9: "sept.", 10: "oct.", 11: "nov.", 12: "déc."}
_MOIS_INSCR = {1: "janv", 2: "févr", 3: "mars", 4: "avr", 5: "mai", 6: "juin",
               7: "juil", 8: "août", 9: "sept", 10: "oct", 11: "nov", 12: "déc"}


class ExcelValidationError(Exception):
    """Erreurs de validation bloquantes (liste de messages lisibles)."""

    def __init__(self, errors: list[str]) -> None:
        self.errors = errors
        super().__init__(f"{len(errors)} erreur(s) de validation")


# --- Helpers ----------------------------------------------------------------
def _r(x: float) -> int:
    """Arrondi « half-up » (identique à Math.round du front)."""
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _fr(value: float, dec: int = 2) -> str:
    """Nombre à `dec` décimales, virgule française."""
    return f"{value:.{dec}f}".replace(".", ",")


def _fr_sep(n: int) -> str:
    """Séparateur de milliers = espace insécable fine (comme le front)."""
    return f"{n:,}".replace(",", " ")


def _pct_signe(value: float, dec: int = 1) -> str:
    """Variation en % signée : +9,2 % / −3,0 %."""
    s = f"{value:+.{dec}f}".replace(".", ",")
    return s + " %"


def _label_mois(d: _dt.date) -> str:
    """Libellé court d'un mois : juil.25, août 25, mars 26…"""
    m = _MOIS_COURTS[d.month]
    yy = f"{d.year % 100:02d}"
    # 'juil.25' (pas d'espace après le point) vs 'août 25' / 'mars 26' (espace).
    return f"{m}{yy}" if m.endswith(".") else f"{m} {yy}"


def _label_mois_bandeau(d: _dt.date) -> str:
    """Libellé du bandeau d'en-tête : toujours un espace avant l'année.

    Ex. 'juil. 25', 'juin 26' (différent des labels de graphe compacts).
    """
    return f"{_MOIS_COURTS[d.month]} {d.year % 100:02d}"


def _mois_suivant(d: _dt.date) -> _dt.date:
    return _dt.date(d.year + (d.month // 12), (d.month % 12) + 1, 1)


def _as_date(v) -> _dt.date | None:
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    return None


def _is_num(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# --- Parsing ----------------------------------------------------------------
def _read_sheet(wb, name: str, errors: list[str]) -> list[dict]:
    """Lit un onglet en liste de dicts après contrôle des en-têtes."""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        errors.append(f"[{name}] onglet vide")
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    expected = SCHEMA.get(name) or OPTIONAL_SCHEMA[name]
    # On tolère des colonnes vides en fin de ligne d'en-tête (artefact Excel).
    while header and header[-1] == "":
        header.pop()
    if header != expected:
        errors.append(
            f"[{name}] colonnes invalides.\n"
            f"      attendu : {expected}\n"
            f"      trouvé  : {header}"
        )
        return []
    out = []
    for i, raw in enumerate(rows[1:], start=2):
        # Ligne entièrement vide -> ignorée (padding Excel).
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue
        rec = {col: raw[j] if j < len(raw) else None for j, col in enumerate(expected)}
        rec["_ligne"] = i
        out.append(rec)
    return out


def _check_sheets(wb, errors: list[str]) -> bool:
    manquants = [s for s in SCHEMA if s not in wb.sheetnames]
    if manquants:
        errors.append(f"Onglet(s) manquant(s) : {', '.join(manquants)}")
    connus = set(SCHEMA) | set(OPTIONAL_SCHEMA) | IGNORED_SHEETS
    inconnus = [s for s in wb.sheetnames if s not in connus]
    if inconnus:
        errors.append(f"Onglet(s) inconnu(s) : {', '.join(inconnus)} "
                      f"(attendus : {', '.join(SCHEMA)} ; "
                      f"facultatifs : {', '.join(OPTIONAL_SCHEMA)} ; "
                      f"ignorés : {', '.join(sorted(IGNORED_SHEETS))})")
    return not manquants


def _require_num(rec: dict, col: str, sheet: str, errors: list[str],
                 *, allow_none: bool = False) -> None:
    v = rec.get(col)
    if v is None and allow_none:
        return
    if not _is_num(v):
        errors.append(f"[{sheet}] ligne {rec['_ligne']}, colonne '{col}' : "
                      f"nombre attendu, trouvé {v!r}")


def _require_neg(rec: dict, col: str, sheet: str, errors: list[str]) -> None:
    v = rec.get(col)
    if _is_num(v) and v > 0:
        errors.append(f"[{sheet}] ligne {rec['_ligne']}, colonne '{col}' : "
                      f"charge attendue en négatif, trouvé {v}")


# --- Validation -------------------------------------------------------------
def _validate(data: dict[str, list[dict]], errors: list[str]) -> None:
    # --- parametres
    params = {str(r["cle"]).strip(): r["valeur"] for r in data["parametres"]
              if r.get("cle") is not None}
    for p in REQUIRED_PARAMS:
        if p not in params:
            errors.append(f"[parametres] clé obligatoire manquante : '{p}'")
    for p in ("ltm_debut", "ltm_fin", "date_maj"):
        if p in params and _as_date(params[p]) is None:
            errors.append(f"[parametres] '{p}' doit être une date, trouvé {params[p]!r}")

    # --- mensuel_groupe : 24 mois consécutifs
    mg = data["mensuel_groupe"]
    if len(mg) != 24:
        errors.append(f"[mensuel_groupe] 24 mois attendus (12 LTM N-1 + 12 LTM N), "
                      f"trouvé {len(mg)}")
    for r in mg:
        if _as_date(r["mois"]) is None:
            errors.append(f"[mensuel_groupe] ligne {r['_ligne']} : 'mois' doit être une date")
        for c in SCHEMA["mensuel_groupe"][1:]:
            _require_num(r, c, "mensuel_groupe", errors)
    mois = [_as_date(r["mois"]) for r in mg if _as_date(r["mois"])]
    if len(mois) == len(mg) and mois:
        for i in range(1, len(mois)):
            if mois[i] != _mois_suivant(mois[i - 1]):
                errors.append(
                    f"[mensuel_groupe] mois non consécutifs : {mois[i-1]:%m/%Y} "
                    f"suivi de {mois[i]:%m/%Y} (ligne {mg[i]['_ligne']})")
                break
        if len(mois) == 24 and "ltm_debut" in params:
            deb = _as_date(params["ltm_debut"])
            if deb and mois[12] != deb:
                errors.append(f"[mensuel_groupe] la ligne 14 ({mois[12]:%m/%Y}) doit "
                              f"correspondre à parametres.ltm_debut ({deb:%m/%Y})")
        if len(mois) == 24 and "ltm_fin" in params:
            fin = _as_date(params["ltm_fin"])
            if fin and mois[23] != fin:
                errors.append(f"[mensuel_groupe] la ligne 25 ({mois[23]:%m/%Y}) doit "
                              f"correspondre à parametres.ltm_fin ({fin:%m/%Y})")

    # --- pnl_ltm : postes exacts, signes
    attendus = PNL_CA_POSTES + PNL_CHARGE_POSTES
    postes = [str(r["poste"]).strip() for r in data["pnl_ltm"] if r.get("poste")]
    if postes != attendus:
        errors.append(f"[pnl_ltm] postes invalides.\n      attendu : {attendus}\n"
                      f"      trouvé  : {postes}")
    for r in data["pnl_ltm"]:
        _require_num(r, "montant_keur", "pnl_ltm", errors)
        _require_num(r, "montant_n1_keur", "pnl_ltm", errors)
        poste = str(r.get("poste") or "").strip()
        if poste in PNL_CHARGE_POSTES:
            _require_neg(r, "montant_keur", "pnl_ltm", errors)
            _require_neg(r, "montant_n1_keur", "pnl_ltm", errors)
        elif poste in PNL_CA_POSTES and _is_num(r["montant_keur"]) and r["montant_keur"] < 0:
            errors.append(f"[pnl_ltm] ligne {r['_ligne']} : le poste CA '{poste}' "
                          f"doit être positif")

    # --- tresorerie_13s
    for r in data["tresorerie_13s"]:
        if not str(r.get("semaine") or "").strip():
            errors.append(f"[tresorerie_13s] ligne {r['_ligne']} : 'semaine' vide")
        _require_num(r, "solde_reel_meur", "tresorerie_13s", errors, allow_none=True)
        _require_num(r, "solde_projete_meur", "tresorerie_13s", errors, allow_none=True)

    # --- covenants
    for r in data["covenants"]:
        _require_num(r, "valeur", "covenants", errors)
        _require_num(r, "seuil", "covenants", errors)
        if str(r.get("sens") or "").strip() not in ("<", ">"):
            errors.append(f"[covenants] ligne {r['_ligne']} : 'sens' doit valoir "
                          f"'<' ou '>', trouvé {r.get('sens')!r}")

    # --- echeances_lbo
    for r in data["echeances_lbo"]:
        if _as_date(r["date"]) is None:
            errors.append(f"[echeances_lbo] ligne {r['_ligne']} : 'date' doit être une date")
        _require_num(r, "montant_keur", "echeances_lbo", errors, allow_none=True)

    # --- inscriptions / filieres / acquisition
    for r in data["inscriptions"]:
        if _as_date(r["mois"]) is None:
            errors.append(f"[inscriptions] ligne {r['_ligne']} : 'mois' doit être une date")
        _require_num(r, "rentree_n", "inscriptions", errors, allow_none=True)
        _require_num(r, "rentree_n1", "inscriptions", errors, allow_none=True)
    for r in data["filieres_groupe"]:
        _require_num(r, "etudiants", "filieres_groupe", errors)

    # --- campus : référentiel
    campus_ids = [str(r["campus_id"]).strip() for r in data["campus"] if r.get("campus_id")]
    if not campus_ids:
        errors.append("[campus] aucun campus")
    if len(set(campus_ids)) != len(campus_ids):
        errors.append("[campus] campus_id en double")
    for r in data["campus"]:
        for c in ["capacite", "etudiants", "pct_alternance", "reussite_bts_pct",
                  "effectif_moyen_classe", "attrition_pct", "nps", "cout_prof_keur",
                  "loyer_annuel_keur", "surface_m2"]:
            _require_num(r, c, "campus", errors)
        if _is_num(r["capacite"]) and r["capacite"] <= 0:
            errors.append(f"[campus] ligne {r['_ligne']} : 'capacite' doit être > 0")

    ref = set(campus_ids)
    # --- campus_mensuel : mêmes campus_id, 12 mois LTM chacun
    mois_ltm = mois[12:] if len(mois) == 24 else []
    par_campus: dict[str, list] = {}
    for r in data["campus_mensuel"]:
        cid = str(r.get("campus_id") or "").strip()
        if cid not in ref:
            errors.append(f"[campus_mensuel] ligne {r['_ligne']} : campus_id '{cid}' "
                          f"absent de l'onglet campus")
        _require_num(r, "ca_keur", "campus_mensuel", errors)
        if _as_date(r["mois"]) is None:
            errors.append(f"[campus_mensuel] ligne {r['_ligne']} : 'mois' doit être une date")
        else:
            par_campus.setdefault(cid, []).append(_as_date(r["mois"]))
    if mois_ltm:
        for cid in ref:
            got = sorted(par_campus.get(cid, []))
            if got != mois_ltm:
                errors.append(f"[campus_mensuel] campus '{cid}' : les 12 mois LTM "
                              f"({mois_ltm[0]:%m/%Y}→{mois_ltm[-1]:%m/%Y}) sont attendus, "
                              f"trouvé {len(got)} mois")

    # --- campus_pnl : mêmes campus_id, signes
    ids_pnl = [str(r.get("campus_id") or "").strip() for r in data["campus_pnl"]]
    if set(ids_pnl) != ref:
        errors.append(f"[campus_pnl] campus_id différents de l'onglet campus : "
                      f"manquants={sorted(ref - set(ids_pnl))}, "
                      f"en trop={sorted(set(ids_pnl) - ref)}")
    for r in data["campus_pnl"]:
        for c in CAMPUS_CA_COLS:
            _require_num(r, c, "campus_pnl", errors)
        for c in CAMPUS_CHARGE_COLS:
            _require_num(r, c, "campus_pnl", errors)
            _require_neg(r, c, "campus_pnl", errors)

    # --- previsionnel : 3 scénarios complets
    par_scen: dict[str, list] = {}
    for r in data["previsionnel"]:
        s = str(r.get("scenario") or "").strip()
        if s not in SCENARIOS:
            errors.append(f"[previsionnel] ligne {r['_ligne']} : scénario '{s}' inconnu "
                          f"(attendus : {', '.join(SCENARIOS)})")
            continue
        if _as_date(r["mois"]) is None:
            errors.append(f"[previsionnel] ligne {r['_ligne']} : 'mois' doit être une date")
        for c in ["ca_keur", "ebitda_keur", "tresorerie_meur"]:
            _require_num(r, c, "previsionnel", errors)
        par_scen.setdefault(s, []).append(_as_date(r["mois"]))
    for s in SCENARIOS:
        got = par_scen.get(s, [])
        if len(got) != 12:
            errors.append(f"[previsionnel] scénario '{s}' : 12 mois attendus, "
                          f"trouvé {len(got)}")
    if len(par_scen) == 3:
        refs = sorted(par_scen["base"])
        for s in ("bas", "haut"):
            if sorted(par_scen[s]) != refs:
                errors.append(f"[previsionnel] scénario '{s}' : mois différents du "
                              f"scénario 'base'")

    # --- hypotheses
    hyp_scen = {str(r.get("scenario") or "").strip() for r in data["hypotheses"]}
    for s in SCENARIOS:
        if s not in hyp_scen:
            errors.append(f"[hypotheses] aucune hypothèse pour le scénario '{s}'")
    for r in data["hypotheses"]:
        _require_num(r, "ordre", "hypotheses", errors)

    # --- points_cles
    for r in data["points_cles"]:
        _require_num(r, "numero", "points_cles", errors)
        if _as_date(r["mois"]) is None:
            errors.append(f"[points_cles] ligne {r['_ligne']} : 'mois' doit être une date")

    # --- campus_filieres (facultatif) : campus_id connus + total = effectif campus
    if data.get("campus_filieres"):
        etu_ref = {str(r["campus_id"]).strip(): r["etudiants"] for r in data["campus"]}
        par_c: dict[str, float] = {}
        for r in data["campus_filieres"]:
            cid = str(r.get("campus_id") or "").strip()
            if cid not in ref:
                errors.append(f"[campus_filieres] ligne {r['_ligne']} : campus_id "
                              f"'{cid}' absent de l'onglet campus")
            _require_num(r, "etudiants", "campus_filieres", errors)
            if not str(r.get("filiere") or "").strip():
                errors.append(f"[campus_filieres] ligne {r['_ligne']} : 'filiere' vide")
            if _is_num(r.get("etudiants")):
                par_c[cid] = par_c.get(cid, 0) + r["etudiants"]
        for cid, total in par_c.items():
            if cid in etu_ref and _is_num(etu_ref[cid]) and abs(total - etu_ref[cid]) > 0:
                errors.append(f"[campus_filieres] campus '{cid}' : somme des filières "
                              f"({total:g}) ≠ effectif du référentiel ({etu_ref[cid]:g})")


def _controles_coherence(data: dict[str, list[dict]], errors: list[str]) -> None:
    """Contrôles de cohérence recalculés en Python (tolérance TOLERANCE_KEUR).

    On ne lit JAMAIS l'onglet 'controles' du fichier : on refait les calculs.
    """
    mg = data["mensuel_groupe"]
    if len(mg) != 24:
        return  # déjà signalé, inutile d'empiler des erreurs dérivées
    n = mg[12:]
    ca_mensuel = sum(r["ca_reel_keur"] for r in n)
    eb_mensuel = sum(r["ebitda_reel_keur"] for r in n)

    pnl = {str(r["poste"]).strip(): r["montant_keur"] for r in data["pnl_ltm"]}
    if all(p in pnl for p in PNL_CA_POSTES + PNL_CHARGE_POSTES):
        ca_pnl = sum(pnl[p] for p in PNL_CA_POSTES)
        eb_pnl = ca_pnl + sum(pnl[p] for p in PNL_CHARGE_POSTES)
        _ecart("CA groupe LTM : mensuel vs P&L", ca_mensuel, ca_pnl, errors)
        _ecart("EBITDA groupe LTM : mensuel vs P&L", eb_mensuel, eb_pnl, errors)

    cp = data["campus_pnl"]
    if cp:
        ca_campus = sum(sum(r[c] for c in CAMPUS_CA_COLS if _is_num(r[c])) for r in cp)
        eb_campus = sum(sum(r[c] for c in CAMPUS_CA_COLS + CAMPUS_CHARGE_COLS
                            if _is_num(r[c])) for r in cp)
        _ecart("CA groupe LTM : mensuel vs somme campus", ca_mensuel, ca_campus, errors)
        _ecart("EBITDA groupe LTM : mensuel vs somme campus", eb_mensuel, eb_campus, errors)

    etu_campus = sum(r["etudiants"] for r in data["campus"] if _is_num(r["etudiants"]))
    etu_fil = sum(r["etudiants"] for r in data["filieres_groupe"] if _is_num(r["etudiants"]))
    _ecart("Étudiants : somme campus vs somme filières", etu_campus, etu_fil, errors,
           unite="")

    if cp and data["campus_mensuel"]:
        ca_ref = sum(sum(r[c] for c in CAMPUS_CA_COLS if _is_num(r[c])) for r in cp)
        ca_men = sum(r["ca_keur"] for r in data["campus_mensuel"] if _is_num(r["ca_keur"]))
        _ecart("CA campus : référentiel vs somme campus_mensuel", ca_ref, ca_men, errors)


def _ecart(libelle: str, a: float, b: float, errors: list[str], unite: str = " k€") -> None:
    if abs(a - b) > TOLERANCE_KEUR:
        errors.append(f"[cohérence] {libelle} : {a:,.0f}{unite} vs {b:,.0f}{unite} "
                      f"(écart {a - b:+,.0f}{unite}, tolérance {TOLERANCE_KEUR}{unite})"
                      .replace(",", " "))


# --- Dérivation vers la structure du dashboard ------------------------------
def _build_dashboard(data: dict[str, list[dict]]) -> dict:
    """Transforme les tables validées en dict aligné sur get_dashboard_data()."""
    params = {str(r["cle"]).strip(): r["valeur"] for r in data["parametres"]}
    date_maj = _as_date(params["date_maj"])
    ltm_debut, ltm_fin = _as_date(params["ltm_debut"]), _as_date(params["ltm_fin"])

    mg = data["mensuel_groupe"]
    n1, n = mg[:12], mg[12:]

    # --- Agrégats groupe (tout est calculé, rien n'est saisi) ---
    ca = sum(r["ca_reel_keur"] for r in n)
    ca_n1 = sum(r["ca_reel_keur"] for r in n1)
    eb = sum(r["ebitda_reel_keur"] for r in n)
    eb_n1 = sum(r["ebitda_reel_keur"] for r in n1)
    budget_ca = sum(r["ca_budget_keur"] for r in n)
    budget_eb = sum(r["ebitda_budget_keur"] for r in n)
    marge = (eb / ca * 100) if ca else 0.0

    tres = [r["tresorerie_brute_meur"] for r in n]
    tres_fin = tres[-1]
    tres_moy12 = sum(tres) / len(tres)
    emprunts_fin = n[-1]["emprunts_meur"]
    emprunts_n1 = n1[-1]["emprunts_meur"]          # juin N-1 vs juin N
    creances_fin = n[-1]["creances_opco_meur"]
    dso_fin, dso_n1 = n[-1]["dso_opco_jours"], n1[-1]["dso_opco_jours"]

    # Effectif : somme des campus. La référence N-1 disponible dans le fichier est
    # le dernier point du cumul d'inscriptions de la rentrée N-1 (onglet
    # inscriptions) : c'est la seule base de comparaison fournie par le contrat.
    etudiants = sum(r["etudiants"] for r in data["campus"])
    etu_n1_final = next((r["rentree_n1"] for r in reversed(data["inscriptions"])
                         if _is_num(r["rentree_n1"])), None)

    def _croissance(a: float, b: float) -> str:
        return _pct_signe((a / b - 1) * 100) if b else ""

    # --- KPIs (6) ---
    kpis = [
        {"lbl": "Chiffre d'affaires LTM", "val": _fr(ca / 1000), "unit": "M€",
         "sub": "vs LTM N-1", "chip": _croissance(ca, ca_n1), "chip_cls": "up"},
        {"lbl": "EBITDA LTM", "val": _fr(eb / 1000), "unit": "M€",
         "sub": f"marge {_fr(marge, 1)} %", "chip": _croissance(eb, eb_n1),
         "chip_cls": "up"},
        {"lbl": "Trésorerie brute", "val": _fr(tres_fin), "unit": "M€",
         "sub": "vs moy. 12M", "chip": _delta_meur(tres_fin - tres_moy12),
         "chip_cls": "up"},
        {"lbl": "Emprunts", "val": _fr(emprunts_fin), "unit": "M€",
         "sub": "vs N-1", "chip": _delta_meur(emprunts_fin - emprunts_n1),
         "chip_cls": "up"},
        {"lbl": "Étudiants inscrits", "val": _fr_sep(etudiants), "unit": "",
         "sub": "vs N-1",
         "chip": _croissance(etudiants, etu_n1_final) if etu_n1_final else "",
         "chip_cls": "up"},
        {"lbl": "Créances OPCO", "val": _fr(creances_fin), "unit": "M€",
         "sub": f"DSO {_r(dso_fin)} j", "chip": f"{_r(dso_fin - dso_n1):+d} j",
         "chip_cls": "warn"},
    ]

    # --- P&L LTM groupe ---
    # Le CA total et sa croissance suivent le mensuel (source de vérité du groupe),
    # pour rester cohérents avec le KPI. Le détail par poste et sa variation N-1
    # viennent du P&L (colonnes montant_n1_keur).
    pnl = {str(r["poste"]).strip(): r for r in data["pnl_ltm"]}
    pnl_table = [{"label": "Chiffre d'affaires", "val": ca,
                  "delta": _croissance(ca, ca_n1), "delta_cls": "d-up", "cls": ""}]
    for p in PNL_CA_POSTES:
        r = pnl[p]
        pnl_table.append({"label": PNL_LABELS[p], "val": r["montant_keur"],
                          "delta": _croissance(r["montant_keur"], r["montant_n1_keur"]),
                          "delta_cls": "d-up", "cls": "sub"})
    # Coûts pédagogiques puis marge brute, comme le gabarit v4.
    ped = pnl["couts_pedagogiques"]
    pnl_table.append({"label": PNL_LABELS["couts_pedagogiques"], "val": ped["montant_keur"],
                      "delta": _croissance(abs(ped["montant_keur"]),
                                           abs(ped["montant_n1_keur"])),
                      "delta_cls": "d-down", "cls": ""})
    mb = ca + ped["montant_keur"]
    mb_n1 = ca_n1 + ped["montant_n1_keur"]
    pnl_table.append({"label": "Marge brute", "pct": f"{_fr(mb / ca * 100, 1)} %",
                      "val": mb, "delta": _croissance(mb, mb_n1),
                      "delta_cls": "d-up", "cls": "tot"})
    for p in PNL_CHARGE_POSTES[1:]:
        r = pnl[p]
        var = _croissance(abs(r["montant_keur"]), abs(r["montant_n1_keur"]))
        # Une charge qui progresse plus vite que le CA est signalée en rouge.
        cls = "d-down" if _hausse_forte(r, ca, ca_n1) else ""
        pnl_table.append({"label": PNL_LABELS[p], "val": r["montant_keur"],
                          "delta": var, "delta_cls": cls, "cls": ""})
    pnl_table.append({"label": "EBITDA", "pct": f"{_fr(marge, 1)} %", "val": eb,
                      "delta": _croissance(eb, eb_n1), "delta_cls": "d-up", "cls": "tot"})

    # --- Covenants : statut calculé depuis valeur/seuil/sens ---
    covenants = []
    for r in data["covenants"]:
        sens = str(r["sens"]).strip()
        ok = r["valeur"] < r["seuil"] if sens == "<" else r["valeur"] > r["seuil"]
        covenants.append({
            "name": str(r["nom"]),
            "value": f"{_fr(r['valeur'], 1)}×",
            "rule": f"doit être {sens} {_fr(r['seuil'], 1)}×",
            "status": "ok" if ok else "ko",
        })

    # --- Échéances : montant vide => badge doc ---
    echeances = []
    for r in data["echeances_lbo"]:
        d = _as_date(r["date"])
        e = {"label": str(r["libelle"]), "date": f"{d:%d/%m}"}
        if _is_num(r["montant_keur"]):
            e["amount"] = f"{_fr_sep(_r(r['montant_keur']))} k€"
        else:
            e["doc"] = True
        echeances.append(e)

    # --- Campus ---
    cpnl = {str(r["campus_id"]).strip(): r for r in data["campus_pnl"]}
    cmens: dict[str, dict] = {}
    for r in data["campus_mensuel"]:
        cmens.setdefault(str(r["campus_id"]).strip(), {})[_as_date(r["mois"])] = r["ca_keur"]
    mois_ltm = [_as_date(r["mois"]) for r in n]

    # Filières par campus : uniquement si l'onglet facultatif est fourni.
    cfil: dict[str, list[tuple[str, int]]] = {}
    for r in data.get("campus_filieres") or []:
        cfil.setdefault(str(r["campus_id"]).strip(), []).append(
            (str(r["filiere"]), _r(r["etudiants"])))

    campus = []
    for r in data["campus"]:
        cid = str(r["campus_id"]).strip()
        p = cpnl[cid]
        c_ca = sum(p[c] for c in CAMPUS_CA_COLS)
        c_eb = sum(p[c] for c in CAMPUS_CA_COLS + CAMPUS_CHARGE_COLS)
        c_ped = p["couts_pedagogiques"]
        c_mb = c_ca + c_ped
        pnl_rows = [{"label": "Chiffre d'affaires", "val": c_ca, "cls": ""}]
        pnl_rows += [{"label": CAMPUS_PNL_LABELS[c], "val": p[c], "cls": "sub"}
                     for c in CAMPUS_CA_COLS]
        pnl_rows.append({"label": CAMPUS_PNL_LABELS["couts_pedagogiques"],
                         "val": c_ped, "cls": ""})
        pnl_rows.append({"label": "Marge brute", "val": c_mb, "cls": "tot",
                         "pct": f"{_fr(c_mb / c_ca * 100, 1)} %" if c_ca else ""})
        pnl_rows += [{"label": CAMPUS_PNL_LABELS[c], "val": p[c], "cls": ""}
                     for c in CAMPUS_CHARGE_COLS[1:]]
        pnl_rows.append({"label": "EBITDA", "val": c_eb, "cls": "tot",
                         "pct": f"{_fr(c_eb / c_ca * 100, 1)} %" if c_ca else ""})

        nom = str(r["nom"])
        campus.append({
            "key": cid,
            "nom": nom,
            "court": nom.split(" — ")[-1],
            "adr": str(r["adresse"]),
            "cap": _r(r["capacite"]),
            "et": _r(r["etudiants"]),
            "alt": _r(r["pct_alternance"]),
            "classe": _r(r["effectif_moyen_classe"]),
            "reu": _r(r["reussite_bts_pct"]),
            "att": r["attrition_pct"],
            "nps": _r(r["nps"]),
            "loyer": _r(r["loyer_annuel_keur"]),
            "m2": _r(r["surface_m2"]),
            "prof": _r(r["cout_prof_keur"]),
            "ca": c_ca,
            "eb": c_eb,
            # Dérivés (jamais saisis)
            "fill": _r(r["etudiants"] / r["capacite"] * 100),
            "alternants": _r(r["etudiants"] * r["pct_alternance"] / 100),
            "marge": f"{_fr(c_eb / c_ca * 100, 1)} %" if c_ca else "",
            "ca_m": [_r(cmens[cid][m]) for m in mois_ltm],
            # Filières du campus : renseignées seulement si l'onglet facultatif
            # campus_filieres est fourni. Sinon listes vides -> le template masque
            # la carte (on n'invente aucune répartition).
            "fil_labels": [f[0] for f in cfil.get(cid, [])],
            "fil_data": [f[1] for f in cfil.get(cid, [])],
            "pnl": pnl_rows,
        })

    # --- Prévisionnel ---
    mois_prev = sorted({_as_date(r["mois"]) for r in data["previsionnel"]})
    scen: dict[str, dict] = {}
    for src, dst in SCENARIOS.items():
        lignes = {_as_date(r["mois"]): r for r in data["previsionnel"]
                  if str(r["scenario"]).strip() == src}
        scen[dst] = {
            "ca": [_r(lignes[m]["ca_keur"]) for m in mois_prev],
            "eb": [_r(lignes[m]["ebitda_keur"]) for m in mois_prev],
            "cash": [lignes[m]["tresorerie_meur"] for m in mois_prev],
            "hyp": [_emphase(str(h["texte"])) for h in sorted(
                (h for h in data["hypotheses"] if str(h["scenario"]).strip() == src),
                key=lambda h: h["ordre"])],
        }

    # points_cles : le mois est résolu en index du graphe de trésorerie projetée.
    cash_pts = []
    for r in sorted(data["points_cles"], key=lambda r: r["numero"]):
        m = _as_date(r["mois"])
        if m in mois_prev:
            cash_pts.append({"i": mois_prev.index(m), "n": str(_r(r["numero"]))})

    # --- Assemblage final (structure identique à demo_data.build()) ---
    from demo_data import GRAY, VIOLET, VIOLET_SOFT, gap_text

    fil = data["filieres_groupe"]
    # Les filières « BTS » restent violettes, les autres en violet clair
    # (même règle visuelle que le mockup).
    fil_colors = [VIOLET if str(f["filiere"]).startswith("BTS") else VIOLET_SOFT
                  for f in fil]

    return {
        "meta": {
            "titre": "Vue globale",
            "exercice": str(params["exercice_libelle"]),
            "ltm": f"{_label_mois_bandeau(ltm_debut)} → {_label_mois_bandeau(ltm_fin)}",
            "maj": f"{date_maj:%d/%m/%Y}",
            "sync": f"{date_maj:%d/%m/%Y}",
            "footer": f"Source : fichier Excel importé · données au {date_maj:%d/%m/%Y}",
        },
        "kpis": kpis,
        "pnl_note": "k€ · cumulé LTM vs budget",
        "pnl_table": pnl_table,
        "covenants": covenants,
        "echeances": echeances,
        "acquisition": [
            {"label": str(r["indicateur"]), "ltm": str(r["valeur"]),
             "n1": str(r["variation_n1"]),
             "n1_cls": "d-down" if str(r["variation_n1"]).startswith("-")
                       or "€" in str(r["variation_n1"]) and "+" in str(r["variation_n1"])
                       and "CAC" in str(r["indicateur"]) else "d-up"}
            for r in data["acquisition"]
        ],
        "campus": campus,
        # Couleurs des barres « filières » du détail campus : même règle que le
        # groupe (BTS en violet, le reste en violet clair), calées sur le nombre
        # réel de filières du premier campus renseigné.
        "campus_fil_colors": [
            VIOLET if str(lbl).startswith("BTS") else VIOLET_SOFT
            for lbl in (next((c["fil_labels"] for c in campus if c["fil_labels"]), []))
        ],
        "charts": {
            "l12": [_label_mois(m) for m in mois_ltm],
            "ca_m": [_r(r["ca_reel_keur"]) for r in n],
            "eb_m": [_r(r["ebitda_reel_keur"]) for r in n],
            "budget_ca": _r(budget_ca),
            "budget_eb": _r(budget_eb),
            "gap_ca": gap_text(_r(ca) - _r(budget_ca)),
            "gap_eb": gap_text(_r(eb) - _r(budget_eb)),
            "budget_label": "Budget N",
            "leg_reel": "Réalisé cumulé",
            "leg_budget": "Budget",
            "cash": {
                "labels": [str(r["semaine"]) for r in data["tresorerie_13s"]],
                "realise": [r["solde_reel_meur"] for r in data["tresorerie_13s"]],
                "projete": [r["solde_projete_meur"] for r in data["tresorerie_13s"]],
                "seuil": [_seuil_covenant(data)] * len(data["tresorerie_13s"]),
                "leg_realise": "Réalisé",
                "leg_projete": "Projeté",
                "seuil_label": f"Seuil covenant {_fr(_seuil_covenant(data), 1)} M€",
            },
            "inscr": {
                "labels": [_MOIS_INSCR[_as_date(r["mois"]).month]
                           for r in data["inscriptions"]],
                "d2026": [r["rentree_n"] for r in data["inscriptions"]],
                "d2025": [r["rentree_n1"] for r in data["inscriptions"]],
                "leg_2026": f"Rentrée {ltm_fin.year}",
                "leg_2025": f"Rentrée {ltm_fin.year - 1}",
                "color_2026": VIOLET,
                "color_2025": GRAY,
            },
            "fil": {
                "labels": [str(f["filiere"]) for f in fil],
                "data": [_r(f["etudiants"]) for f in fil],
                "colors": fil_colors,
            },
        },
        "projections": {
            "labels": [_label_mois(m) for m in mois_prev],
            "scenarios": [{"key": SCENARIOS[s], "label": SCENARIO_LABELS[SCENARIOS[s]]}
                          for s in ("base", "bas", "haut")],
            "seuil": [_seuil_covenant(data)] * len(mois_prev),
            "seuil_label": f"Seuil {_fr(_seuil_covenant(data), 1)} M€",
            "cash_pts": cash_pts,
            "leg_ca": "CA",
            "leg_eb": "EBITDA",
            "leg_cash": "Trésorerie",
            "scen": scen,
        },
        "points_cles": [_emphase(str(r["texte"])) for r in
                        sorted(data["points_cles"], key=lambda r: r["numero"])],
        "demo_blocks": [],
    }


def _hausse_forte(row: dict, ca: float, ca_n1: float) -> bool:
    """Charge dont la croissance dépasse celle du CA -> signalée en rouge."""
    if not (_is_num(row["montant_n1_keur"]) and row["montant_n1_keur"]):
        return False
    var_charge = abs(row["montant_keur"]) / abs(row["montant_n1_keur"]) - 1
    var_ca = (ca / ca_n1 - 1) if ca_n1 else 0
    return var_charge > var_ca


def _seuil_covenant(data: dict[str, list[dict]]) -> float:
    """Seuil de trésorerie affiché sur les graphes (M€).

    Le fichier ne porte pas de seuil de trésorerie dédié : on retient 1,0 M€,
    valeur du plancher de liquidité suivi par la direction.
    """
    return 1.0


# Mise en gras des textes libres (hypothèses / points clés). Le fichier fournit
# du texte brut ; on rebalise pour retrouver le rendu du mockup.
# Règle : seule la valeur qui suit un « : » est mise en valeur. Une valeur entre
# parenthèses (« (+6 %) ») ou sans « : » (« Attrition : 4,8 % » -> si, ici) reste
# gérée par ce seul motif, ce qui évite d'empiler des cas particuliers.
_ESPACES = r"\s  "          # espace, insécable, insécable fine
_BOLD_APRES_DEUX_POINTS = _re.compile(
    r"(:\s*)("
    r"\(?\d[\d" + _ESPACES + r"]*[.,]?\d*\)?\s*M€"        # (1,5) M€ · 0,6 M€
    r"|[+−-]?\d+[.,]?\d*\s*%"                              # +2 %
    r"|\d[\d" + _ESPACES + r"]*\s+étudiants"               # 2 990 étudiants
    r"|\d+(?![\d" + _ESPACES + r"]*[%€])"                  # nombre isolé : « campus : 1 »
    r")"
)


def _emphase(texte: str) -> str:
    """Encadre de <b>…</b> la valeur qui suit un « : » dans un texte libre.

    Idempotent : un texte déjà balisé n'est pas retouché.
    """
    if "<b>" in texte:
        return texte
    return _BOLD_APRES_DEUX_POINTS.sub(r"\1<b>\2</b>", texte)


def _delta_meur(diff: float) -> str:
    """Variation en M€ : positif '+0,3 M€', négatif '(0,9) M€' (convention front)."""
    if diff < 0:
        return f"({_fr(abs(diff), 1)}) M€"
    return f"+{_fr(diff, 1)} M€"


# --- API publique -----------------------------------------------------------
def parse_and_validate(content: bytes) -> tuple[dict, dict[str, list[dict]]]:
    """Parse + valide un classeur en mémoire.

    Retourne (dashboard, tables_brutes). Lève ExcelValidationError si invalide.
    """
    errors: list[str] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelValidationError([f"Fichier illisible : {type(exc).__name__} — {exc}"])

    try:
        if not _check_sheets(wb, errors):
            raise ExcelValidationError(errors)
        data = {name: _read_sheet(wb, name, errors) for name in SCHEMA}
        # Onglets facultatifs : lus seulement s'ils sont présents.
        for name in OPTIONAL_SCHEMA:
            data[name] = _read_sheet(wb, name, errors) if name in wb.sheetnames else []
        if errors:
            raise ExcelValidationError(errors)
        _validate(data, errors)
        if errors:
            raise ExcelValidationError(errors)
        _controles_coherence(data, errors)
        if errors:
            raise ExcelValidationError(errors)
        return _build_dashboard(data), data
    finally:
        wb.close()


def file_hash(content: bytes) -> str:
    """Empreinte SHA256 du fichier importé (traçabilité)."""
    return hashlib.sha256(content).hexdigest()


def volumetrie(tables: dict[str, list[dict]]) -> dict[str, int]:
    """Nombre de lignes par onglet (historisation / diff)."""
    return {name: len(rows) for name, rows in tables.items()}
